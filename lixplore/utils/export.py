#!/usr/bin/env python3

"""
Export utilities for Lixplore - support multiple academic formats
"""

import csv
import json
import os
from datetime import datetime
from typing import List, Dict
import xml.etree.ElementTree as ET
from xml.dom import minidom

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    XLSX_AVAILABLE = True
except ImportError:
    XLSX_AVAILABLE = False


# Default export directory within the project
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
DEFAULT_EXPORT_DIR = os.path.join(PROJECT_ROOT, "exports")

# Define all export format folders (one folder per format type)
EXPORT_FOLDERS = {
    'csv': 'csv',
    'json': 'json',
    'bibtex': 'bibtex',
    'ris': 'ris',
    'endnote': 'endnote_xml',        # EndNote XML format
    'enw': 'endnote_tagged',          # EndNote tagged format (.enw)
    'xlsx': 'excel',                  # Excel format (.xlsx)
    'xml': 'xml'                      # Generic XML format
}


def initialize_export_folders():
    """
    Create all export folders on initialization.
    Called once to ensure all folders exist.
    """
    if not os.path.exists(DEFAULT_EXPORT_DIR):
        os.makedirs(DEFAULT_EXPORT_DIR)
        print(f"✓ Created main export directory: {DEFAULT_EXPORT_DIR}")
    
    # Create all format-specific subfolders
    for format_key, folder_name in EXPORT_FOLDERS.items():
        folder_path = os.path.join(DEFAULT_EXPORT_DIR, folder_name)
        if not os.path.exists(folder_path):
            os.makedirs(folder_path)
    
    # Create README in exports folder
    readme_path = os.path.join(DEFAULT_EXPORT_DIR, "README.md")
    if not os.path.exists(readme_path):
        with open(readme_path, 'w', encoding='utf-8') as f:
            f.write("# Lixplore Exports Directory\n\n")
            f.write("This directory contains all exported search results organized by format.\n\n")
            f.write("## Export Folders:\n\n")
            f.write("- **bibtex/** - BibTeX format (.bib) for LaTeX citations\n")
            f.write("- **csv/** - CSV format (.csv) for Excel, Google Sheets\n")
            f.write("- **endnote_tagged/** - EndNote Tagged format (.enw) - native EndNote import\n")
            f.write("- **endnote_xml/** - EndNote XML format (.xml) - EndNote XML import\n")
            f.write("- **excel/** - Microsoft Excel format (.xlsx)\n")
            f.write("- **json/** - JSON format (.json) - structured data\n")
            f.write("- **ris/** - RIS format (.ris) - reference managers (Zotero, Mendeley)\n")
            f.write("- **xml/** - Generic XML format (.xml)\n\n")
            f.write("## Usage:\n\n")
            f.write("All exports are automatically saved to their designated folder:\n\n")
            f.write("```bash\n")
            f.write("lixplore -P -q \"your query\" -m 10 -e csv      # → csv/\n")
            f.write("lixplore -P -q \"your query\" -m 10 -e xlsx     # → excel/\n")
            f.write("lixplore -P -q \"your query\" -m 10 -e enw      # → endnote_tagged/\n")
            f.write("lixplore -P -q \"your query\" -m 10 -e endnote  # → endnote_xml/\n")
            f.write("```\n\n")
            f.write("Files are automatically organized - no manual management needed!\n")


def get_export_directory(format: str = None) -> str:
    """
    Get the export directory for a specific format.
    
    Args:
        format: Export format (csv, json, etc.)
    
    Returns:
        Path to export directory
    """
    # Initialize folders if they don't exist
    if not os.path.exists(DEFAULT_EXPORT_DIR):
        initialize_export_folders()
    
    # Return format-specific folder
    if format and format in EXPORT_FOLDERS:
        folder_path = os.path.join(DEFAULT_EXPORT_DIR, EXPORT_FOLDERS[format])
        # Ensure the folder exists (in case it was deleted)
        if not os.path.exists(folder_path):
            os.makedirs(folder_path)
        return folder_path
    
    return DEFAULT_EXPORT_DIR


def generate_filename(base_name: str, extension: str, format: str = None) -> str:
    """
    Generate a filename with timestamp in the appropriate export directory.
    
    Args:
        base_name: Base name for the file
        extension: File extension
        format: Export format (for subfolder organization)
    
    Returns:
        Full path to the file
    """
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{base_name}_{timestamp}.{extension}"
    
    # Get the appropriate directory
    export_dir = get_export_directory(format)
    
    return os.path.join(export_dir, filename)


def export_to_csv(results: List[Dict], filename: str = None) -> str:
    """
    Export results to CSV format.
    
    Args:
        results: List of article dictionaries
        filename: Output filename (optional)
    
    Returns:
        Path to exported file
    """
    if not results:
        print("No results to export.")
        return None
    
    if not filename:
        filename = generate_filename("lixplore_results", "csv", "csv")
    elif not os.path.isabs(filename):
        # If relative path provided, put it in csv subfolder
        filename = os.path.join(get_export_directory("csv"), filename)
    
    # Define CSV columns
    fieldnames = [
        "title", "authors", "abstract", "journal", "year", 
        "doi", "url", "source"
    ]
    
    with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=fieldnames, extrasaction='ignore')
        writer.writeheader()
        
        for result in results:
            # Convert authors list to string
            row = result.copy()
            if isinstance(row.get('authors'), list):
                row['authors'] = "; ".join(row['authors'])
            writer.writerow(row)
    
    print(f"✓ Exported {len(results)} results to: {filename}")
    return filename


def export_to_json(results: List[Dict], filename: str = None) -> str:
    """
    Export results to JSON format.
    
    Args:
        results: List of article dictionaries
        filename: Output filename (optional)
    
    Returns:
        Path to exported file
    """
    if not results:
        print("No results to export.")
        return None
    
    if not filename:
        filename = generate_filename("lixplore_results", "json", "json")
    elif not os.path.isabs(filename):
        # If relative path provided, put it in json subfolder
        filename = os.path.join(get_export_directory("json"), filename)
    
    with open(filename, 'w', encoding='utf-8') as jsonfile:
        json.dump(results, jsonfile, indent=2, ensure_ascii=False)
    
    print(f"✓ Exported {len(results)} results to: {filename}")
    return filename


def export_to_bibtex(results: List[Dict], filename: str = None) -> str:
    """
    Export results to BibTeX format.
    
    Args:
        results: List of article dictionaries
        filename: Output filename (optional)
    
    Returns:
        Path to exported file
    """
    if not results:
        print("No results to export.")
        return None
    
    if not filename:
        filename = generate_filename("lixplore_results", "bib", "bibtex")
    elif not os.path.isabs(filename):
        # If relative path provided, put it in bibtex subfolder
        filename = os.path.join(get_export_directory("bibtex"), filename)
    
    with open(filename, 'w', encoding='utf-8') as bibfile:
        for i, result in enumerate(results, start=1):
            # Generate citation key
            first_author = ""
            if result.get('authors'):
                first_author = result['authors'][0].split()[-1] if result['authors'] else ""
            year = result.get('year', datetime.now().year)
            citation_key = f"{first_author}{year}_{i}" if first_author else f"article{year}_{i}"
            
            # Determine entry type
            entry_type = "article"
            
            # Start BibTeX entry
            bibfile.write(f"@{entry_type}{{{citation_key},\n")
            
            # Add fields
            if result.get('title'):
                title = result['title'].replace('{', '').replace('}', '')
                bibfile.write(f"  title = {{{title}}},\n")
            
            if result.get('authors'):
                authors = " and ".join(result['authors'])
                bibfile.write(f"  author = {{{authors}}},\n")
            
            if result.get('journal'):
                bibfile.write(f"  journal = {{{result['journal']}}},\n")
            
            if result.get('year'):
                bibfile.write(f"  year = {{{result['year']}}},\n")
            
            if result.get('doi'):
                bibfile.write(f"  doi = {{{result['doi']}}},\n")
            
            if result.get('url'):
                bibfile.write(f"  url = {{{result['url']}}},\n")
            
            if result.get('abstract'):
                abstract = result['abstract'].replace('{', '').replace('}', '')
                bibfile.write(f"  abstract = {{{abstract}}},\n")
            
            # Close entry
            bibfile.write("}\n\n")
    
    print(f"✓ Exported {len(results)} results to: {filename}")
    return filename


def export_to_ris(results: List[Dict], filename: str = None) -> str:
    """
    Export results to RIS format (Research Information Systems).
    
    Args:
        results: List of article dictionaries
        filename: Output filename (optional)
    
    Returns:
        Path to exported file
    """
    if not results:
        print("No results to export.")
        return None
    
    if not filename:
        filename = generate_filename("lixplore_results", "ris", "ris")
    elif not os.path.isabs(filename):
        # If relative path provided, put it in ris subfolder
        filename = os.path.join(get_export_directory("ris"), filename)
    
    with open(filename, 'w', encoding='utf-8') as risfile:
        for result in results:
            # TY - Type of reference (JOUR = Journal Article)
            risfile.write("TY  - JOUR\n")
            
            # TI - Title
            if result.get('title'):
                risfile.write(f"TI  - {result['title']}\n")
            
            # AU - Authors (one per line)
            if result.get('authors'):
                for author in result['authors']:
                    risfile.write(f"AU  - {author}\n")
            
            # JO - Journal name
            if result.get('journal'):
                risfile.write(f"JO  - {result['journal']}\n")
            
            # PY - Publication year
            if result.get('year'):
                risfile.write(f"PY  - {result['year']}\n")
            
            # DO - DOI
            if result.get('doi'):
                risfile.write(f"DO  - {result['doi']}\n")
            
            # UR - URL
            if result.get('url'):
                risfile.write(f"UR  - {result['url']}\n")
            
            # AB - Abstract
            if result.get('abstract'):
                risfile.write(f"AB  - {result['abstract']}\n")
            
            # DB - Database (source)
            if result.get('source'):
                risfile.write(f"DB  - {result['source']}\n")
            
            # ER - End of reference
            risfile.write("ER  - \n\n")
    
    print(f"✓ Exported {len(results)} results to: {filename}")
    return filename


def export_to_xlsx(results: List[Dict], filename: str = None) -> str:
    """
    Export results to Excel (XLSX) format.
    
    Args:
        results: List of article dictionaries
        filename: Output filename (optional)
    
    Returns:
        Path to exported file
    """
    if not XLSX_AVAILABLE:
        print("Error: openpyxl is not installed. Install it with: pip install openpyxl")
        return None
    
    if not results:
        print("No results to export.")
        return None
    
    if not filename:
        filename = generate_filename("lixplore_results", "xlsx", "xlsx")
    elif not os.path.isabs(filename):
        # If relative path provided, put it in xlsx subfolder
        filename = os.path.join(get_export_directory("xlsx"), filename)
    
    # Create workbook and active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Search Results"
    
    # Define headers
    headers = ["#", "Title", "Authors", "Journal", "Year", "DOI", "URL", "Source", "Abstract"]
    
    # Style for header row
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Write data
    for row_idx, result in enumerate(results, start=2):
        ws.cell(row=row_idx, column=1, value=row_idx-1)  # Number
        ws.cell(row=row_idx, column=2, value=result.get('title', ''))
        
        # Authors as comma-separated string
        authors = result.get('authors', [])
        authors_str = ", ".join(authors) if isinstance(authors, list) else str(authors)
        ws.cell(row=row_idx, column=3, value=authors_str)
        
        ws.cell(row=row_idx, column=4, value=result.get('journal', ''))
        ws.cell(row=row_idx, column=5, value=result.get('year', ''))
        ws.cell(row=row_idx, column=6, value=result.get('doi', ''))
        ws.cell(row=row_idx, column=7, value=result.get('url', ''))
        ws.cell(row=row_idx, column=8, value=result.get('source', ''))
        ws.cell(row=row_idx, column=9, value=result.get('abstract', ''))
    
    # Adjust column widths
    column_widths = [5, 50, 30, 30, 8, 25, 40, 12, 80]
    for col_idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[chr(64 + col_idx)].width = width
    
    # Enable text wrapping for all cells
    for row in ws.iter_rows(min_row=2, max_row=len(results)+1):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    
    # Save workbook
    wb.save(filename)
    
    print(f"✓ Exported {len(results)} results to: {filename}")
    return filename


def export_to_endnote_xml(results: List[Dict], filename: str = None) -> str:
    """
    Export results to EndNote XML format.
    
    Args:
        results: List of article dictionaries
        filename: Output filename (optional)
    
    Returns:
        Path to exported file
    """
    if not results:
        print("No results to export.")
        return None
    
    if not filename:
        filename = generate_filename("lixplore_results_endnote", "xml", "endnote")
    elif not os.path.isabs(filename):
        # If relative path provided, put it in endnote subfolder
        filename = os.path.join(get_export_directory("endnote"), filename)
    
    # Create XML structure
    root = ET.Element("xml")
    records = ET.SubElement(root, "records")
    
    for i, result in enumerate(results, start=1):
        record = ET.SubElement(records, "record")
        
        # Database
        database = ET.SubElement(record, "database", name="lixplore", path="lixplore.xml")
        database.text = "lixplore"
        
        # Source type - Journal Article
        source_type = ET.SubElement(record, "source-type", name="Journal Article")
        source_type.text = "17"
        
        # Record number
        rec_number = ET.SubElement(record, "rec-number")
        rec_number.text = str(i)
        
        # Foreign keys
        foreign_keys = ET.SubElement(record, "foreign-keys")
        key = ET.SubElement(foreign_keys, "key", app="EN", db_id="0")
        key.text = str(i)
        
        # Reference type
        ref_type = ET.SubElement(record, "ref-type", name="Journal Article")
        ref_type.text = "17"
        
        # Contributors (Authors)
        if result.get('authors'):
            contributors = ET.SubElement(record, "contributors")
            authors = ET.SubElement(contributors, "authors")
            for author in result['authors']:
                author_elem = ET.SubElement(authors, "author")
                author_elem.text = author
        
        # Titles
        titles = ET.SubElement(record, "titles")
        if result.get('title'):
            title = ET.SubElement(titles, "title")
            title.text = result['title']
        
        # Periodical (Journal)
        if result.get('journal'):
            periodical = ET.SubElement(record, "periodical")
            full_title = ET.SubElement(periodical, "full-title")
            full_title.text = result['journal']
        
        # Dates
        if result.get('year'):
            dates = ET.SubElement(record, "dates")
            year = ET.SubElement(dates, "year")
            year.text = str(result['year'])
        
        # URLs
        if result.get('url'):
            urls = ET.SubElement(record, "urls")
            related_urls = ET.SubElement(urls, "related-urls")
            url = ET.SubElement(related_urls, "url")
            url.text = result['url']
        
        # Electronic Resource Number (DOI)
        if result.get('doi'):
            electronic_resource_num = ET.SubElement(record, "electronic-resource-num")
            electronic_resource_num.text = result['doi']
        
        # Abstract
        if result.get('abstract'):
            abstract = ET.SubElement(record, "abstract")
            abstract.text = result['abstract']
        
        # Custom fields
        custom1 = ET.SubElement(record, "custom1")
        custom1.text = result.get('source', '')
    
    # Pretty print XML
    xml_str = minidom.parseString(ET.tostring(root)).toprettyxml(indent="  ")
    
    with open(filename, 'w', encoding='utf-8') as f:
        f.write(xml_str)
    
    print(f"✓ Exported {len(results)} results to EndNote XML: {filename}")
    return filename


def export_to_enw(results: List[Dict], filename: str = None) -> str:
    """
    Export results to EndNote Tagged Format (.enw).
    This is the native EndNote import format using text tags.
    
    Args:
        results: List of article dictionaries
        filename: Output filename (optional)
    
    Returns:
        Path to exported file
    """
    if not results:
        print("No results to export.")
        return None
    
    if not filename:
        filename = generate_filename("lixplore_results", "enw", "enw")
    elif not os.path.isabs(filename):
        # If relative path provided, put it in enw subfolder
        filename = os.path.join(get_export_directory("enw"), filename)
    
    with open(filename, 'w', encoding='utf-8') as enwfile:
        for result in results:
            # %0 - Type of reference (Journal Article)
            enwfile.write("%0 Journal Article\n")
            
            # %T - Title
            if result.get('title'):
                enwfile.write(f"%T {result['title']}\n")
            
            # %A - Authors (one per line)
            if result.get('authors'):
                for author in result['authors']:
                    enwfile.write(f"%A {author}\n")
            
            # %J - Journal name
            if result.get('journal'):
                enwfile.write(f"%J {result['journal']}\n")
            
            # %D - Publication year
            if result.get('year'):
                enwfile.write(f"%D {result['year']}\n")
            
            # %R - DOI
            if result.get('doi'):
                enwfile.write(f"%R {result['doi']}\n")
            
            # %U - URL
            if result.get('url'):
                enwfile.write(f"%U {result['url']}\n")
            
            # %X - Abstract
            if result.get('abstract'):
                enwfile.write(f"%X {result['abstract']}\n")
            
            # %~ - Name of database (source)
            if result.get('source'):
                enwfile.write(f"%~ {result['source']}\n")
            
            # End of record (blank line)
            enwfile.write("\n")
    
    print(f"✓ Exported {len(results)} results to EndNote format: {filename}")
    return filename


def export_to_xml(results: List[Dict], filename: str = None) -> str:
    """
    Export results to generic XML format.
    
    Args:
        results: List of article dictionaries
        filename: Output filename (optional)
    
    Returns:
        Path to exported file
    """
    if not results:
        print("No results to export.")
        return None
    
    if not filename:
        filename = generate_filename("lixplore_results", "xml", "xml")
    elif not os.path.isabs(filename):
        # If relative path provided, put it in xml subfolder
        filename = os.path.join(get_export_directory("xml"), filename)
    
    # Create XML structure
    root = ET.Element("search_results")
    root.set("count", str(len(results)))
    root.set("exported_at", datetime.now().isoformat())
    
    for i, result in enumerate(results, start=1):
        article = ET.SubElement(root, "article")
        article.set("id", str(i))
        
        # Add all fields
        for key, value in result.items():
            if value:  # Only add non-empty values
                elem = ET.SubElement(article, key)
                if isinstance(value, list):
                    # Handle lists (like authors)
                    for item in value:
                        sub_elem = ET.SubElement(elem, "item")
                        sub_elem.text = str(item)
                else:
                    elem.text = str(value)
    
    # Pretty print XML
    xml_str = minidom.parseString(ET.tostring(root)).toprettyxml(indent="  ")
    
    with open(filename, 'w', encoding='utf-8') as f:
        f.write(xml_str)
    
    print(f"✓ Exported {len(results)} results to XML: {filename}")
    return filename


def export_results(results: List[Dict], format: str, filename: str = None) -> str:
    """
    Main export function that routes to appropriate exporter.
    
    Args:
        results: List of article dictionaries
        format: Export format ('csv', 'json', 'bibtex', 'ris', 'endnote', 'enw', 'xlsx', 'xml')
        filename: Output filename (optional)
    
    Returns:
        Path to exported file
    """
    format = format.lower()
    
    if format == 'csv':
        return export_to_csv(results, filename)
    elif format == 'json':
        return export_to_json(results, filename)
    elif format == 'bibtex':
        return export_to_bibtex(results, filename)
    elif format == 'ris':
        return export_to_ris(results, filename)
    elif format == 'xlsx':
        return export_to_xlsx(results, filename)
    elif format == 'endnote':
        return export_to_endnote_xml(results, filename)
    elif format == 'enw':
        return export_to_enw(results, filename)
    elif format == 'xml':
        return export_to_xml(results, filename)
    else:
        print(f"Error: Unsupported export format '{format}'")
        return None
